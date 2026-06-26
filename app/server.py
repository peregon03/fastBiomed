from __future__ import annotations

import base64
import csv
import io
import json
import os
import sqlite3
from datetime import date, datetime
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from urllib.parse import parse_qs, urlparse


ROOT = Path(__file__).resolve().parent
STATIC_DIR = ROOT / "static"
DB_PATH = Path(os.environ.get("BIOMED_DB_PATH", ROOT / "data" / "biomed.db"))
SERVER_HOST = os.environ.get("BIOMED_HOST", "127.0.0.1")
SERVER_PORT = int(os.environ.get("BIOMED_PORT", "8000"))
SEED_DEMO = os.environ.get("BIOMED_SEED_DEMO", "1").lower() in ("1", "true", "yes", "si")
AREAS = ("UCI", "Urgencias", "Cirugia")


def today() -> date:
    return date.today()


def parse_date(value: str | None) -> date | None:
    if not value:
        return None
    return datetime.strptime(value[:10], "%Y-%m-%d").date()


def add_months(source: date, months: int) -> date:
    month = source.month - 1 + months
    year = source.year + month // 12
    month = month % 12 + 1
    days = [31, 29 if year % 4 == 0 and (year % 100 != 0 or year % 400 == 0) else 28, 31, 30, 31, 30,
            31, 31, 30, 31, 30, 31]
    return source.replace(year=year, month=month, day=min(source.day, days[month - 1]))


def status_for(next_maintenance: str | None, pending: int = 0) -> str:
    next_date = parse_date(next_maintenance)
    if not next_date:
        return "Sin programacion"
    delta = (next_date - today()).days
    if delta < 0:
        return "Vencido"
    if pending:
        return "Pendiente"
    if delta <= 30:
        return "Proximo a vencer"
    return "Vigente"


def connect() -> sqlite3.Connection:
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON")
    return conn


def init_db() -> None:
    DB_PATH.parent.mkdir(parents=True, exist_ok=True)
    with connect() as conn:
        conn.executescript(
            """
            CREATE TABLE IF NOT EXISTS equipment (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL,
                plate TEXT NOT NULL UNIQUE,
                serial_number TEXT NOT NULL,
                location TEXT NOT NULL,
                area TEXT NOT NULL CHECK(area IN ('UCI', 'Urgencias', 'Cirugia')),
                brand TEXT,
                model TEXT,
                specific_location TEXT,
                requires_maintenance INTEGER NOT NULL DEFAULT 1,
                last_maintenance TEXT,
                next_maintenance TEXT,
                frequency_months INTEGER NOT NULL DEFAULT 6,
                pending_intervention INTEGER NOT NULL DEFAULT 0,
                created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                updated_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
            );

            CREATE TABLE IF NOT EXISTS maintenance_history (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                equipment_id INTEGER NOT NULL REFERENCES equipment(id) ON DELETE CASCADE,
                performed_at TEXT NOT NULL,
                previous_next_date TEXT,
                new_next_date TEXT NOT NULL,
                notes TEXT,
                created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
            );
            """
        )
        migrate_db(conn)
        count = conn.execute("SELECT COUNT(*) FROM equipment").fetchone()[0]
        if count == 0 and SEED_DEMO:
            seed(conn)


def migrate_db(conn: sqlite3.Connection) -> None:
    """Migra bases de datos existentes al esquema actual sin perder datos."""
    existing = {row[1] for row in conn.execute("PRAGMA table_info(equipment)").fetchall()}
    if "requires_maintenance" in existing:
        return  # Ya migrado

    # Reconstruir tabla para eliminar CHECK(frequency_months IN (6,12)) y agregar columnas nuevas.
    # SQLite no permite DROP CONSTRAINT ni MODIFY COLUMN, por eso se usa la estrategia rename/copy/drop.
    conn.executescript("""
        PRAGMA foreign_keys = OFF;

        CREATE TABLE equipment_new (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            plate TEXT NOT NULL UNIQUE,
            serial_number TEXT NOT NULL,
            location TEXT NOT NULL,
            area TEXT NOT NULL CHECK(area IN ('UCI', 'Urgencias', 'Cirugia')),
            brand TEXT,
            model TEXT,
            specific_location TEXT,
            requires_maintenance INTEGER NOT NULL DEFAULT 1,
            last_maintenance TEXT,
            next_maintenance TEXT,
            frequency_months INTEGER NOT NULL DEFAULT 6,
            pending_intervention INTEGER NOT NULL DEFAULT 0,
            created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
            updated_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
        );

        INSERT INTO equipment_new
            (id, name, plate, serial_number, location, area,
             last_maintenance, next_maintenance, frequency_months,
             pending_intervention, created_at, updated_at)
        SELECT
            id, name, plate, serial_number, location, area,
            last_maintenance, next_maintenance, frequency_months,
            pending_intervention, created_at, updated_at
        FROM equipment;

        DROP TABLE equipment;
        ALTER TABLE equipment_new RENAME TO equipment;

        PRAGMA foreign_keys = ON;
    """)


def seed(conn: sqlite3.Connection) -> None:
    # (name, plate, serial_number, location, area, brand, model, specific_location,
    #  requires_maintenance, last_maintenance, next_maintenance, frequency_months, pending_intervention)
    samples = [
        ("Monitor multiparametro", "UCI-001", "SN-MON-9001", "Cubiculo 1", "UCI", "Mindray", "MEC-1200", "Cubículo 1", 1, "2026-02-10", "2026-08-10", 6, 0),
        ("Ventilador mecanico", "UCI-002", "SN-VM-7821", "Cubiculo 3", "UCI", "Drager", "Evita 4", "Cubículo 3", 1, "2025-11-20", "2026-05-20", 6, 1),
        ("Bomba de infusion", "UCI-003", "SN-BI-1220", "Estacion central", "UCI", "Baxter", "Colleague", "Estación central", 1, "2026-01-05", "2027-01-05", 12, 0),
        ("Desfibrilador", "URG-001", "SN-DES-3112", "Sala reanimacion", "Urgencias", "Philips", "HeartStart", "Sala reanimación", 1, "2025-12-01", "2026-06-01", 6, 0),
        ("Electrocardiografo", "URG-002", "SN-ECG-8172", "Consultorio 2", "Urgencias", "GE Healthcare", "MAC 5500", "Consultorio 2", 1, "2026-03-22", "2026-09-22", 6, 0),
        ("Lampara cielitica", "CIR-001", "SN-LC-4400", "Quirofano 1", "Cirugia", "Mopec", "Mopec-500", "Quirófano 1", 1, "2025-09-15", "2026-09-15", 12, 0),
        ("Maquina de anestesia", "CIR-002", "SN-AN-2318", "Quirofano 2", "Cirugia", "Drager", "Fabius GS", "Quirófano 2", 1, None, None, 6, 0),
        ("Mesa quirurgica", "CIR-003", "SN-MQ-6404", "Quirofano 3", "Cirugia", "Trumpf", "Jupiter", "Quirófano 3", 1, "2026-05-30", "2026-11-30", 6, 0),
        ("Flujometro de oxigeno", "UCI-004", "SN-FL-0041", "Cubiculo 4", "UCI", None, None, "Cubículo 4", 0, None, None, 0, 0),
        ("Termometro digital", "URG-003", "SN-TM-0031", "Triaje", "Urgencias", None, None, "Sala de triaje", 0, None, None, 0, 0),
    ]
    conn.executemany(
        """
        INSERT INTO equipment
        (name, plate, serial_number, location, area, brand, model, specific_location,
         requires_maintenance, last_maintenance, next_maintenance, frequency_months, pending_intervention)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        samples,
    )
    history = [
        (1, "2026-02-10", "2026-02-10", "2026-08-10", "Carga inicial"),
        (3, "2026-01-05", "2026-01-05", "2027-01-05", "Carga inicial"),
        (5, "2026-03-22", "2026-03-22", "2026-09-22", "Carga inicial"),
        (8, "2026-05-30", "2026-05-30", "2026-11-30", "Carga inicial"),
    ]
    conn.executemany(
        """
        INSERT INTO maintenance_history
        (equipment_id, performed_at, previous_next_date, new_next_date, notes)
        VALUES (?, ?, ?, ?, ?)
        """,
        history,
    )


def row_to_equipment(row: sqlite3.Row) -> dict:
    data = dict(row)
    data["status"] = status_for(data.get("next_maintenance"), data.get("pending_intervention", 0))
    next_date = parse_date(data.get("next_maintenance"))
    data["days_until_due"] = None if not next_date else (next_date - today()).days
    return data


def json_body(handler: BaseHTTPRequestHandler) -> dict:
    length = int(handler.headers.get("Content-Length", "0"))
    if length == 0:
        return {}
    raw = handler.rfile.read(length).decode("utf-8")
    return json.loads(raw or "{}")


def read_equipment(conn: sqlite3.Connection, query: dict | None = None) -> list[dict]:
    query = query or {}
    sql = "SELECT * FROM equipment"
    args: list[str] = []
    clauses: list[str] = []
    if query.get("area"):
        clauses.append("area = ?")
        args.append(query["area"])
    if clauses:
        sql += " WHERE " + " AND ".join(clauses)
    sql += " ORDER BY area, name"
    rows = [row_to_equipment(r) for r in conn.execute(sql, args).fetchall()]
    if query.get("status"):
        rows = [r for r in rows if r["status"] == query["status"]]
    return rows


def dashboard() -> dict:
    with connect() as conn:
        all_equipment = read_equipment(conn)
        history = conn.execute("SELECT equipment_id, performed_at FROM maintenance_history ORDER BY performed_at").fetchall()

    # KPIs y alertas solo sobre equipos con mantenimiento preventivo
    equipment = [e for e in all_equipment if e.get("requires_maintenance")]
    inventory_only = len(all_equipment) - len(equipment)

    total = len(equipment)
    scheduled = sum(1 for item in equipment if item["next_maintenance"])
    completed_ids = set(h["equipment_id"] for h in history)
    completed = sum(1 for item in equipment if item["id"] in completed_ids)
    overdue = sum(1 for item in equipment if item["status"] == "Vencido")
    due_soon = sum(1 for item in equipment if item["status"] == "Proximo a vencer")
    no_schedule = sum(1 for item in equipment if item["status"] == "Sin programacion")
    pending = sum(1 for item in equipment if item["status"] == "Pendiente")
    denominator = completed + overdue + due_soon + pending
    compliance = round((completed / denominator) * 100, 1) if denominator > 0 else 0

    status_counts: dict[str, int] = {}
    area_counts = {area: 0 for area in AREAS}
    for item in equipment:
        status_counts[item["status"]] = status_counts.get(item["status"], 0) + 1
        area_counts[item["area"]] = area_counts.get(item["area"], 0) + 1

    month_counts: dict[str, int] = {}
    for row in history:
        key = row["performed_at"][:7]
        month_counts[key] = month_counts.get(key, 0) + 1

    alerts = [
        item for item in equipment
        if item["status"] in ("Vencido", "Proximo a vencer", "Sin programacion", "Pendiente")
    ]

    return {
        "totals": {
            "equipment": total,
            "inventoryOnly": inventory_only,
            "scheduled": scheduled,
            "dueSoon": due_soon,
            "overdue": overdue,
            "completed": completed,
            "compliance": compliance,
            "noSchedule": no_schedule,
            "pending": pending,
        },
        "statusCounts": status_counts,
        "areaCounts": area_counts,
        "monthlyCompleted": month_counts,
        "alerts": alerts,
    }


def calendar_items(query: dict) -> list[dict]:
    with connect() as conn:
        items = read_equipment(conn, query)
    return [
        {
            "id": item["id"],
            "title": item["name"],
            "area": item["area"],
            "plate": item["plate"],
            "date": item["next_maintenance"],
            "status": item["status"],
            "location": item["location"],
        }
        for item in items
        if item["next_maintenance"]
    ]


def save_equipment(payload: dict, equipment_id: int | None = None) -> dict:
    requires = payload.get("requires_maintenance")
    requires_maintenance = 0 if str(requires).strip().lower() in ("0", "false", "no", "") and requires is not None and str(requires).strip() != "" else 1
    # Si el campo llega como booleano Python o entero
    if isinstance(requires, bool):
        requires_maintenance = 1 if requires else 0
    elif isinstance(requires, int):
        requires_maintenance = 1 if requires else 0

    fields = {
        "name": payload.get("name", "").strip(),
        "plate": payload.get("plate", "").strip(),
        "serial_number": payload.get("serial_number", "").strip(),
        "location": payload.get("location", "").strip(),
        "area": payload.get("area", "").strip(),
        "brand": (payload.get("brand") or "").strip() or None,
        "model": (payload.get("model") or "").strip() or None,
        "specific_location": (payload.get("specific_location") or "").strip() or None,
        "requires_maintenance": requires_maintenance,
        "last_maintenance": payload.get("last_maintenance") or None,
        "next_maintenance": payload.get("next_maintenance") or None,
        "frequency_months": int(payload.get("frequency_months") or 0) if not requires_maintenance else int(payload.get("frequency_months") or 6),
        "pending_intervention": 1 if payload.get("pending_intervention") else 0,
    }
    missing = [k for k in ("name", "plate", "serial_number", "location", "area") if not fields[k]]
    if missing:
        raise ValueError("Campos obligatorios incompletos: " + ", ".join(missing))
    if fields["area"] not in AREAS:
        raise ValueError("Area no valida")
    if fields["requires_maintenance"] and fields["frequency_months"] not in (6, 12):
        raise ValueError("La frecuencia debe ser de 6 o 12 meses para equipos con mantenimiento")

    with connect() as conn:
        if equipment_id:
            conn.execute(
                """
                UPDATE equipment SET
                name=:name, plate=:plate, serial_number=:serial_number, location=:location, area=:area,
                brand=:brand, model=:model, specific_location=:specific_location,
                requires_maintenance=:requires_maintenance,
                last_maintenance=:last_maintenance, next_maintenance=:next_maintenance,
                frequency_months=:frequency_months, pending_intervention=:pending_intervention,
                updated_at=CURRENT_TIMESTAMP
                WHERE id=:id
                """,
                {**fields, "id": equipment_id},
            )
        else:
            cur = conn.execute(
                """
                INSERT INTO equipment
                (name, plate, serial_number, location, area, brand, model, specific_location,
                 requires_maintenance, last_maintenance, next_maintenance, frequency_months, pending_intervention)
                VALUES (:name, :plate, :serial_number, :location, :area, :brand, :model, :specific_location,
                        :requires_maintenance, :last_maintenance, :next_maintenance,
                        :frequency_months, :pending_intervention)
                """,
                fields,
            )
            equipment_id = cur.lastrowid
        row = conn.execute("SELECT * FROM equipment WHERE id = ?", (equipment_id,)).fetchone()
        return row_to_equipment(row)


def complete_maintenance(equipment_id: int, payload: dict) -> dict:
    performed = parse_date(payload.get("performed_at")) or today()
    notes = payload.get("notes", "")
    with connect() as conn:
        row = conn.execute("SELECT * FROM equipment WHERE id = ?", (equipment_id,)).fetchone()
        if not row:
            raise ValueError("Equipo no encontrado")
        if not row["requires_maintenance"]:
            raise ValueError("Este equipo no tiene mantenimiento preventivo programado")
        if not row["frequency_months"]:
            raise ValueError("El equipo no tiene frecuencia de mantenimiento definida")
        new_next = add_months(performed, int(row["frequency_months"]))
        conn.execute(
            """
            UPDATE equipment
            SET last_maintenance = ?, next_maintenance = ?, pending_intervention = 0, updated_at = CURRENT_TIMESTAMP
            WHERE id = ?
            """,
            (performed.isoformat(), new_next.isoformat(), equipment_id),
        )
        conn.execute(
            """
            INSERT INTO maintenance_history
            (equipment_id, performed_at, previous_next_date, new_next_date, notes)
            VALUES (?, ?, ?, ?, ?)
            """,
            (equipment_id, performed.isoformat(), row["next_maintenance"], new_next.isoformat(), notes),
        )
        updated = conn.execute("SELECT * FROM equipment WHERE id = ?", (equipment_id,)).fetchone()
        return row_to_equipment(updated)


MONTH_COLS = ["enero", "febrero", "marzo", "abril", "mayo", "junio",
              "julio", "agosto", "septiembre", "octubre", "noviembre", "diciembre"]

AREA_ALIASES = {
    "uci": "UCI",
    "urgencias": "Urgencias",
    "urgencia": "Urgencias",
    "cirugia": "Cirugia",
    "cirugía": "Cirugia",
    "cirug\u00eda": "Cirugia",
    "cirujia": "Cirugia",
}

IMPORT_ALIASES = {
    "nombre del equipo": "name",
    "nombre": "name",
    "equipo": "name",
    "placa": "plate",
    "número de serie": "serial_number",
    "numero de serie": "serial_number",
    "serie": "serial_number",
    "serial": "serial_number",
    "ubicación": "specific_location",
    "ubicacion": "specific_location",
    "área": "area",
    "area": "area",
    "marca": "brand",
    "brand": "brand",
    "modelo": "model",
    "model": "model",
    "ultimo mantenimiento": "last_maintenance",
    "fecha del ultimo mantenimiento preventivo": "last_maintenance",
    "proximo mantenimiento": "next_maintenance",
    "fecha del proximo mantenimiento preventivo": "next_maintenance",
    "frecuencia mantenimiento": "frequency_months",
    "frecuencia": "frequency_months",
    "requiere mantenimiento preventivo": "requires_maintenance",
    "requiere mantenimiento": "requires_maintenance",
    "requiere": "requires_maintenance",
}


def _parse_sheet_rows(ws: object, sheet_area: str | None = None) -> list[dict]:
    """Extrae filas de una hoja openpyxl, incluyendo columnas de meses."""
    all_rows = list(ws.iter_rows(values_only=True))
    if not all_rows:
        return []
    headers = [str(c or "").strip().lower() for c in all_rows[0]]
    result = []
    for cells in all_rows[1:]:
        row: dict = {headers[i]: cells[i] for i in range(min(len(headers), len(cells)))}
        if sheet_area:
            row["_sheet_area"] = sheet_area
        # Detectar meses programados (x / X) y derivar next_maintenance
        scheduled_months = [i for i, m in enumerate(MONTH_COLS) if headers.count(m) and str(row.get(m) or "").strip().lower() == "x"]
        row["_scheduled_months"] = scheduled_months
        result.append(row)
    return result


def _normalize_row(raw: dict, row_index: int) -> dict:
    """Convierte una fila cruda al payload de save_equipment."""
    normalized: dict = {}
    for k, v in raw.items():
        key = str(k or "").strip().lower()
        mapped = IMPORT_ALIASES.get(key, key)
        normalized[mapped] = v

    # Área: prioridad sheet_area > columna area
    area_raw = str(raw.get("_sheet_area") or normalized.get("area") or "").strip()
    area = AREA_ALIASES.get(area_raw.lower(), area_raw)

    # requires_maintenance
    req_raw = str(normalized.get("requires_maintenance") or "SI").strip().upper()
    requires = req_raw in ("SI", "S", "1", "TRUE", "YES", "SÍ")

    # frequency_months
    freq_raw = str(normalized.get("frequency_months") or "")
    if not requires:
        frequency = 0
    elif "12" in freq_raw or "anual" in freq_raw.lower():
        frequency = 12
    else:
        frequency = 6

    # next_maintenance desde columnas de mes si no viene explícito
    next_maint = normalized.get("next_maintenance") or None
    last_maint = normalized.get("last_maintenance") or None
    if requires and not next_maint:
        scheduled = raw.get("_scheduled_months", [])
        current_year = date.today().year
        future = [m for m in scheduled if m + 1 >= date.today().month]
        if future:
            month_num = future[0] + 1
            next_maint = f"{current_year}-{month_num:02d}-01"
        elif scheduled:
            month_num = scheduled[-1] + 1
            next_maint = f"{current_year + 1}-{month_num:02d}-01"

    # Limpiar None y strings vacíos
    def clean(v: object) -> str | None:
        s = str(v or "").strip()
        return s or None

    return {
        "name": clean(normalized.get("name")),
        "plate": clean(normalized.get("plate")),
        "serial_number": clean(normalized.get("serial_number")),
        "location": clean(normalized.get("specific_location") or normalized.get("location") or area),
        "specific_location": clean(normalized.get("specific_location")),
        "area": area,
        "brand": clean(normalized.get("brand")),
        "model": clean(normalized.get("model")),
        "requires_maintenance": 1 if requires else 0,
        "last_maintenance": clean(last_maint),
        "next_maintenance": clean(next_maint),
        "frequency_months": frequency,
    }


def import_rows(payload: dict) -> dict:
    filename = payload.get("filename", "").lower()
    content = base64.b64decode(payload.get("content", ""))
    raw_rows: list[dict] = []

    if filename.endswith(".xlsx"):
        try:
            import openpyxl  # type: ignore
        except Exception as exc:
            raise ValueError("Para importar XLSX instale openpyxl o cargue el archivo como CSV.") from exc
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        for sheet_name in wb.sheetnames:
            # Inferir área desde el nombre de la hoja
            sheet_area = AREA_ALIASES.get(sheet_name.strip().lower())
            raw_rows.extend(_parse_sheet_rows(wb[sheet_name], sheet_area))
    else:
        text = content.decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(text))
        raw_rows = [{(k or "").strip().lower(): v for k, v in row.items()} for row in reader]

    imported = 0
    skipped = 0
    errors: list[str] = []

    for index, raw in enumerate(raw_rows, start=2):
        try:
            row_payload = _normalize_row(raw, index)
            # Omitir filas completamente vacías
            if not row_payload["name"] and not row_payload["plate"]:
                skipped += 1
                continue
            if not row_payload["plate"]:
                errors.append(f"Fila {index}: sin placa — omitida")
                skipped += 1
                continue
            save_equipment(row_payload)
            imported += 1
        except Exception as exc:
            errors.append(f"Fila {index}: {exc}")

    return {"imported": imported, "skipped": skipped, "errors": errors}


def excel_report() -> bytes:
    with connect() as conn:
        equipment = read_equipment(conn)
    rows = "".join(
        "<tr>"
        f"<td>{e['name']}</td><td>{e['plate']}</td><td>{e['serial_number']}</td>"
        f"<td>{e.get('brand') or ''}</td><td>{e.get('model') or ''}</td>"
        f"<td>{e.get('specific_location') or e['location']}</td><td>{e['area']}</td>"
        f"<td>{'SI' if e.get('requires_maintenance') else 'NO'}</td>"
        f"<td>{e['last_maintenance'] or ''}</td><td>{e['next_maintenance'] or ''}</td>"
        f"<td>{e['frequency_months'] if e.get('requires_maintenance') else ''}</td><td>{e['status']}</td>"
        "</tr>"
        for e in equipment
    )
    html = f"""
    <html><meta charset="utf-8"><body>
    <table border="1">
    <tr><th>Equipo</th><th>Placa</th><th>Serie</th><th>Marca</th><th>Modelo</th>
    <th>Ubicacion especifica</th><th>Area</th><th>Requiere mantenimiento</th>
    <th>Ultimo mantenimiento</th><th>Proximo mantenimiento</th><th>Frecuencia (meses)</th><th>Estado</th></tr>
    {rows}
    </table></body></html>
    """
    return html.encode("utf-8")


def report_html() -> bytes:
    data = dashboard()
    rows = "".join(
        f"<tr><td>{a['name']}</td><td>{a['plate']}</td><td>{a['area']}</td><td>{a['next_maintenance'] or ''}</td><td>{a['status']}</td></tr>"
        for a in data["alerts"]
    )
    html = f"""
    <!doctype html><html><head><meta charset="utf-8"><title>Reporte de mantenimientos</title>
    <style>body{{font-family:Arial;margin:32px;color:#172033}} table{{border-collapse:collapse;width:100%}}
    th,td{{border:1px solid #ccd3df;padding:8px;text-align:left}} th{{background:#edf2f7}}</style></head>
    <body><h1>Reporte de mantenimientos preventivos</h1>
    <p>Total equipos: {data['totals']['equipment']} | Programados: {data['totals']['scheduled']} |
    Vencidos: {data['totals']['overdue']} | Cumplimiento: {data['totals']['compliance']}%</p>
    <h2>Alertas activas</h2><table><tr><th>Equipo</th><th>Placa</th><th>Area</th><th>Fecha</th><th>Estado</th></tr>{rows}</table>
    <script>window.addEventListener('load', () => setTimeout(() => window.print(), 400));</script></body></html>
    """
    return html.encode("utf-8")


class Handler(BaseHTTPRequestHandler):
    def send_json(self, data: dict | list, status: int = 200) -> None:
        raw = json.dumps(data, ensure_ascii=False).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(raw)))
        self.end_headers()
        self.wfile.write(raw)

    def send_file(self, path: Path) -> None:
        if not path.exists() or not path.is_file():
            self.send_error(404)
            return
        content_type = "text/html; charset=utf-8"
        if path.suffix == ".css":
            content_type = "text/css; charset=utf-8"
        elif path.suffix == ".js":
            content_type = "application/javascript; charset=utf-8"
        raw = path.read_bytes()
        self.send_response(200)
        self.send_header("Content-Type", content_type)
        self.send_header("Content-Length", str(len(raw)))
        self.end_headers()
        self.wfile.write(raw)

    def do_GET(self) -> None:
        try:
            parsed = urlparse(self.path)
            qs = {k: v[0] for k, v in parse_qs(parsed.query).items()}
            if parsed.path == "/":
                return self.send_file(STATIC_DIR / "index.html")
            if parsed.path == "/api/health":
                return self.send_json({"ok": True, "database": str(DB_PATH), "host": SERVER_HOST, "port": SERVER_PORT})
            if parsed.path == "/api/dashboard":
                return self.send_json(dashboard())
            if parsed.path == "/api/equipment":
                with connect() as conn:
                    return self.send_json(read_equipment(conn, qs))
            if parsed.path == "/api/calendar":
                return self.send_json(calendar_items(qs))
            if parsed.path == "/api/export/excel":
                raw = excel_report()
                self.send_response(200)
                self.send_header("Content-Type", "application/vnd.ms-excel; charset=utf-8")
                self.send_header("Content-Disposition", "attachment; filename=mantenimientos_biomedicos.xls")
                self.send_header("Content-Length", str(len(raw)))
                self.end_headers()
                return self.wfile.write(raw)
            if parsed.path == "/api/export/pdf":
                raw = report_html()
                self.send_response(200)
                self.send_header("Content-Type", "text/html; charset=utf-8")
                self.send_header("Content-Length", str(len(raw)))
                self.end_headers()
                return self.wfile.write(raw)
            return self.send_file(STATIC_DIR / parsed.path.lstrip("/"))
        except Exception as exc:
            self.send_json({"error": str(exc)}, 500)

    def do_POST(self) -> None:
        try:
            parsed = urlparse(self.path)
            if parsed.path == "/api/equipment":
                return self.send_json(save_equipment(json_body(self)), 201)
            if parsed.path == "/api/import":
                return self.send_json(import_rows(json_body(self)))
            if parsed.path.endswith("/complete") and parsed.path.startswith("/api/equipment/"):
                equipment_id = int(parsed.path.split("/")[3])
                return self.send_json(complete_maintenance(equipment_id, json_body(self)))
            self.send_error(404)
        except sqlite3.IntegrityError as exc:
            self.send_json({"error": f"Registro duplicado o invalido: {exc}"}, 400)
        except Exception as exc:
            self.send_json({"error": str(exc)}, 400)

    def do_PUT(self) -> None:
        try:
            parsed = urlparse(self.path)
            if parsed.path.startswith("/api/equipment/"):
                equipment_id = int(parsed.path.split("/")[3])
                return self.send_json(save_equipment(json_body(self), equipment_id))
            self.send_error(404)
        except Exception as exc:
            self.send_json({"error": str(exc)}, 400)

    def do_DELETE(self) -> None:
        try:
            parsed = urlparse(self.path)
            if parsed.path.startswith("/api/equipment/"):
                equipment_id = int(parsed.path.split("/")[3])
                with connect() as conn:
                    conn.execute("DELETE FROM equipment WHERE id = ?", (equipment_id,))
                return self.send_json({"ok": True})
            self.send_error(404)
        except Exception as exc:
            self.send_json({"error": str(exc)}, 400)

    def log_message(self, fmt: str, *args: object) -> None:
        return


def run() -> None:
    init_db()
    server = ThreadingHTTPServer((SERVER_HOST, SERVER_PORT), Handler)
    display_host = "127.0.0.1" if SERVER_HOST in ("0.0.0.0", "") else SERVER_HOST
    print(f"Sistema biomedico disponible en http://{display_host}:{SERVER_PORT}", flush=True)
    print(f"Base de datos: {DB_PATH}", flush=True)
    server.serve_forever()


if __name__ == "__main__":
    run()

